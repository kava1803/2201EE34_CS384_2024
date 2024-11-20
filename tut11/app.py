import streamlit as st
import openpyxl as opxl
from openpyxl.utils import get_column_letter
import pandas as pd
from io import BytesIO

class Student:
    def __init__(self, roll_number):
        self.roll_number = roll_number
        self.marks = {}
        self.total_scaled = 0
        self.grade = ""

    def calculate_total_scaled(self, max_marks, weightage):
        total_scaled = 0
        for subject, mark in self.marks.items():
            max_mark = max_marks.get(subject, 0)
            weight = weightage.get(subject, 0)
            if max_mark:
                total_scaled += (mark / max_mark) * weight
        self.total_scaled = total_scaled


def process_excel(input_wb):
    active_wb = input_wb.active
    output_wb = opxl.Workbook()
    roll_sorted_ws = output_wb.active
    roll_sorted_ws.title = "Roll_Sorted"
    grade_sorted_ws = output_wb.create_sheet("Grade_Sorted")

    max_marks = {}
    weightage = {}
    subjects = []
    students = {}

    colCount = active_wb.max_column
    rowCount = active_wb.max_row

    # Read subjects, max_marks and weightage
    for col in range(3, colCount + 1):
        subject = active_wb.cell(row=1, column=col).value
        subjects.append(subject)
        max_marks[subject] = active_wb.cell(row=2, column=col).value
        weightage[subject] = active_wb.cell(row=3, column=col).value

    # Initialize Student objects
    for row in range(4, rowCount + 1):
        roll_number = active_wb.cell(row=row, column=1).value
        student = Student(roll_number)

        for col_index in range(3, colCount + 1):
            subject = subjects[col_index - 3]
            mark = active_wb.cell(row=row, column=col_index).value
            student.marks[subject] = mark

        student.calculate_total_scaled(max_marks, weightage)
        students[roll_number] = student

    # Copy original data to new sheet
    for row in active_wb.iter_rows(values_only=True):
        roll_sorted_ws.append(row)
        grade_sorted_ws.append(row)

    # Add Total Scaled/Grade columns
    colCount += 1
    roll_sorted_ws.cell(row=1, column=colCount).value = 'Total Scaled/100'
    roll_sorted_ws.cell(row=1, column=colCount + 1).value = 'Grade'
    grade_sorted_ws.cell(row=1, column=colCount).value = 'Total Scaled/100'
    grade_sorted_ws.cell(row=1, column=colCount + 1).value = 'Grade'

    # Calculate grades
    total_students = len(students)
    grade_cutoffs = {
        "AA": (0.05 * total_students),
        "AB": (0.15 * total_students),
        "BB": (0.25 * total_students),
        "BC": (0.30 * total_students),
        "CC": (0.15 * total_students),
        "CD": (0.05 * total_students),
        "DD": (0.05 * total_students)
    }

    sorted_students = sorted(students.values(), key=lambda s: s.total_scaled, reverse=True)

    current_index = 0
    for grade, count in grade_cutoffs.items():
        for i in range(round(count)):
            if current_index < total_students:
                sorted_students[current_index].grade = grade
                current_index += 1

    # Populate Roll_Sorted and Grade_Sorted sheets
    for row in range(4, rowCount + 1):
        roll_number = active_wb.cell(row=row, column=1).value
        student = students.get(roll_number)

        if student:
            roll_sorted_ws.cell(row=row, column=colCount).value = student.total_scaled
            roll_sorted_ws.cell(row=row, column=colCount + 1).value = student.grade

    for row_idx, student in enumerate(sorted_students, start=4):
        grade_sorted_ws.cell(row=row_idx, column=1).value = student.roll_number
        for col_idx, subject in enumerate(subjects, start=3):
            grade_sorted_ws.cell(row=row_idx, column=col_idx).value = student.marks.get(subject)

        grade_sorted_ws.cell(row=row_idx, column=colCount).value = student.total_scaled
        grade_sorted_ws.cell(row=row_idx, column=colCount + 1).value = student.grade

    # Formatting columns widths
    padding = 3
    for ws in [roll_sorted_ws, grade_sorted_ws]:
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value) + padding
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length

    # Save the processed Excel file
    output_buffer = BytesIO()
    output_wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer


# --- Streamlit UI ---
def main():
    # Page Layout
    st.set_page_config(page_title="IITP Grader", page_icon="🎓", layout="wide")

    # Title and Welcome Section
    st.markdown("""
    # IITP Grader 🎓
    Welcome to the IITP Grader! This tool helps you calculate scaled marks and grades for students based on their input data.
    Upload your Excel file, and get a detailed report in return.
    """)

    st.markdown("### Upload Your Input Excel File")
    
    # File upload section with instructions
    uploaded_file = st.file_uploader(
        "Choose an Excel file (xlsx)",
        type="xlsx",
        help="Upload the Excel file that contains the student marks and max scores. Ensure the file format is correct."
    )

    if uploaded_file is not None:
        # Display a loading spinner
        with st.spinner('Processing the file...'):
            try:
                # Process the Excel file
                input_wb = opxl.load_workbook(uploaded_file)
                output_buffer = process_excel(input_wb)

                # Success message with download option
                st.success("Processing Complete! Your file is ready.")
                st.download_button(
                    label="Download Processed Excel",
                    data=output_buffer,
                    file_name="Output_Excel.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                # Error message
                st.error(f"Error: {e}")
    
    else:
        # Display instructions if no file is uploaded
        st.markdown("""
        📄 **Steps:**
        1. Upload an Excel file containing student marks.
        2. The system will calculate the total scaled marks and assign grades.
        3. Download the processed file with detailed results.
        """)

    # Footer Section (Optional)
    st.markdown("""
    ---
    ### About this tool:
    This tool is designed to assist with the grading process for students. It takes an Excel file with student marks, calculates the scaled totals, and assigns grades based on predefined cutoffs. It generates a report in Excel format that you can download and use.
    """)

if __name__ == "__main__":
    main()
